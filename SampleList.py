import pandas as pd

import openpyxl

from openpyxl import load_workbook
import os
import re


class SampleListFileNameException(Exception):
    pass


class HeaderNotFoundException(Exception):
    pass


class SampleNameHeaderException(Exception):
    pass



class SampleList(object):

    def __init__(self):
        self.sample_list = None
        return

    @property
    def sample_count(self):
        return len(self.sample_list.index)


    def readFile(self, filename, project_name = None, header_row = None, sample_name_header = None):        

        self.path = os.path.abspath(filename)
        self.project_name = self.parseProjectName() if project_name is None else project_name

         ####  Examine XLS file.  Find row with headers and last sample (first empty after samples)
        wb = load_workbook(self.path, data_only=True)
        sh = wb.worksheets[0]

        namestr = 'sample name'
        numstr = 'sample number'

        self.head_row = None
        self.last_row = None

        ## Find header row by searching for 'sample number' 
        for i in range(1, sh.max_row + 1):
            if self.head_row == None:
                for j in range (1, sh.max_column + 1):
                    if sh.cell(row = i, column=j).value == numstr:
                        self.head_row = i
                        break        
        ## if no header row found, throw exception
        if self.head_row == None:
            raise HeaderNotFoundException


        ## make sure name header is correct or throw exception
        if namestr not in [cell.value for cell in sh[self.head_row]]:
            raise SampleNameHeaderException            

        ## find last sample row (assume first empty row is end)
        for i in range(self.head_row, sh.max_row + 1):
            if all([cell.value is None for cell in sh[i]]):
                self.last_row = i
                break

        if self.last_row == None:
            self.last_row = sh.max_row + 1


        ## Read the data from the excel file with Pandas
        data = pd.read_excel(self.path, engine='openpyxl', skiprows=self.head_row - 1, nrows=(self.last_row - self.head_row))

        ## Create new data frame with only the columns we want
        self.sample_list = pd.DataFrame(data=data[[numstr, namestr]].values, columns=['number', 'name'])
        self.sample_list['method'] = "None"
        self.sample_list['position'] = "NA"
        ## Sanitize the names to remove offensive characters
        self.sample_list['name'] = self.sample_list['name'].apply(self.sanitizeName)

        ## add sample number to sample name to conform to Bioinformatics code
        for i in self.sample_list.index:
            name = self.sample_list.loc[i, 'name']

            if type(self.sample_list.loc[i, 'number']) == int:
                name = name + "_sample_{:02d}".format(self.sample_list.loc[i, 'number'])
            else:
                name = name + "_sample_{}".format(self.sample_list.loc[i, 'number'])
            self.sample_list.at[i, 'name'] = name


        return

    ### Remove offensive characters that will mess up file names or Bioinformatics code
    def sanitizeName(self, name):
        newName = re.sub(r"[ +\[\]\.\+!/@#\$%\^&\*\(\)\?\|\\;:]+", '_', str(name))
        if newName[0].isdigit():
            newName = 'S' + newName
        return newName


    ### get the project name from the sample list file name
    def parseProjectName(self):

        split_name = self.path.split('\\')[-1]
        proj_pattern = r'([^/]+_\d{6}.*)(?=_SampleList.xlsx)'
        match = re.match(proj_pattern, split_name)
        if match:
            name = match.group(1)
        else:
            raise SampleListFileNameException       

        return name

    ### get entry from sample table
      #  data - header name as string
      #  index - index to sample list
    def getSampleData(self, data, index):
        return self.sample_list[data][index]


        

