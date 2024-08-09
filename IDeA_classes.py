import os
import sys
import openpyxl
import random
import glob 
import shutil
import re

class Project:
  def __init__(self, project_name = "Test_041223"):
    self.project_name = project_name

class Experiment(Project):
  def __init__(self, species, sample_type, **kwargs):
    self.__dict__.update(kwargs)
    self.species = species
    self.sample_type = sample_type

class Sample(Experiment):
  def __init__(self, **kwargs):
    self.__dict__.update(kwargs)

class Plate(Experiment):
  def __init__(self, index = 0, tray = "G", **kwargs):
    self.__dict__.update(kwargs)
    self.tray = tray
    self.index = index
    self.position = self.index_to_well()
  
  def index_to_well(self): 
    index = self.index
    row = int(index/12)
    row_letter = "ABCDEFGH"[row]
    col = str((index % 12) + 1)
    return(row_letter + col)

class SampleList():
  def __init__(self, path, **kwargs):
    self.abs_path = os.path.abspath(path)
    self.basename = os.path.basename(path)
    self.dirname = os.path.dirname(path)
    self.sample_list_name = "".join(self.basename.rsplit("_SampleList.xlsx"))
    self.path = path
    if not hasattr(self, 'project_name'):
      self.project_name = "_".join(self.basename.split("_")[0:2])
    wb_obj = openpyxl.load_workbook(path)
    self.sheet = wb_obj.active
    self.samplename_cell = self.find_sample_name()
    self.samplenames = self.get_sample_name_list()
  
  def find_species(self):
    x = self.sheet
    for i in range(1, x.max_row + 1):
      row = [x.value for x in x[i]]
      if "organism" in row:
        col = [i for i,x in enumerate(row) if x == "organism"][0] #Index, change to Excel_index
        # print(f'"organism" Cell for {self.project_name} is in position {"ABCDEFGH"[col]}{i}: row {i}, column {col + 1}')
        # print(self.sheet.cell(row = i + 1, column = col + 1).value)
        return(self.sheet.cell(row = i + 1, column = col + 1).value)

    
  def find_sample_name(self):
    x = self.sheet
    for i in range(1, x.max_row + 1):
      row = [x.value for x in x[i]]
      if "sample name" in row:
        col = [i for i,x in enumerate(row) if x == "sample name"][0] #Index, change to Excel_index
        #print(f'"Sample name" Cell for {self.project_name} is in position {"ABCDEFGH"[col]}{i}: row {i}, column {col + 1}')
        return(i, col + 1)
  
  def get_sample_name_list(self):
    i = self.samplename_cell[0] + 1
    j = self.samplename_cell[1]
    w = []
    while True:
      val = self.sheet.cell(row = i, column = j).value
      if val is None:
        break
      if val[0].isdigit():
        val = "S" + val
      w.append(re.sub(r"[ +\[\]]+", '_', val))
      i +=1
    return(w)
  
  def get_position(self, index):
    row = int(index/12)
    row_letter = "ABCDEFGH"[row]
    col = str((index % 12) + 1)
    return(row_letter + col)





class InjectionSequence():
  def __init__(self, path, project, sample_names, tray, start_index, pool_well = None):
    self.path = path
    self.project_name = project
    self.samples = sample_names
    self.tray = tray
    if pool_well is None:
      self.pool_well = "H12"
    else:
      self.pool_well = pool_well
    self.start_index = int(start_index)
    self.positions = [self.get_position(x) for x in range(self.start_index, self.start_index + len(self.samples))]
    self.position_label = (self.get_position(x) for x in range(self.start_index + len(self.samples)))
    self.DIA_path = "C:\\Xcalibur\\methods\\DIA\\60min_5ulLoad_10ulLoop\\DIA_12mz_15k_60min_5ulLoad_10ulLoop_031221"
    self.GPF_path = "C:\\Xcalibur\\methods\\DIA\\60min_5ulLoad_10ulLoop\\DIA_GPF_{mz_start}_{mz_end}_60min_5ulLoad_10ulLoop_031221"
    self.sample_rows = self._pack_samples()
    self.GPF_rows = self._pack_GPFs()
    self.pool_rows = self._pack_pool_runs()
    
  
  def get_position(self, index):
    row = int(index/12)
    row_letter = "ABCDEFGH"[row]
    col = str((index % 12) + 1)
    return(row_letter + col)
  
  def _format_filename(self, project_name, sample_number, sample_name):
    return(f'{project_name}_Sample_{sample_number:02d}_{sample_name}')
  
  def _format_row(self, project_name, sample_name, method, tray, position):
    return(f'Unknown,{sample_name},1,D:\\{project_name},{method},,,{tray}{position},10,,0,0,0,1,,,,,,,\n')

  def _pack_samples(self):
    w = []
    for i,sample in enumerate(self.samples):
      filename = self._format_filename(self.project_name, i + 1, sample)
      w.append(self._format_row(self.project_name, filename, self.DIA_path, self.tray, self.positions[i]))
    return(w)
  
  def _pack_GPFs(self):
    w = []
    for i in range(1,7):
      mz_start = 300 + (100*i)
      mz_end = 400 + (100*i)
      filename = f'{self.project_name}_GPF6_{i}'
      f_method = self.GPF_path.format(mz_start = mz_start, mz_end = mz_end)
      w.append(self._format_row(self.project_name, filename, f_method, self.tray, self.pool_well))
    return(w)
  
  def _pack_pool_runs(self):
    w = []
    for i in range(1,4):
      filename = f'{self.project_name}_Pool_{i}'
      w.append(self._format_row(self.project_name, filename, self.DIA_path, self.tray, self.pool_well))
    return(w)
  
  def write_injection_sequence(self, tail = "", dest = ""):
    line1 = 'Bracket Type=4,,,,,,,,,,,,,,,,,,,,\n'
    line2 = 'Sample Type,File Name,Sample ID,Path,Instrument Method,Process Method,Calibration File,Position,Inj Vol,Level,Sample Wt,Sample Vol,ISTD Amt,Dil Factor,L1 Study,L2 Client,L3 Laboratory,L4 Company,L5 Phone,Comment,Sample Name\n'
    data = list(tuple(self.sample_rows)) #Deep copy
    GPF = self.GPF_rows
    pool = self.pool_rows
    half_index = int(len(data)/2)
    pool_index1 = int(len(data)/3)
    pool_index2 = int(len(data)*2/3)
    for i in range(5, -1, -1):
      data.insert(half_index, GPF[i])
    data.insert(pool_index1, pool[0])
    data.insert(int(len(data)*2/3), pool[1])
    data.append(pool[2])
    if not dest == "":
      dest = dest + "/"
    else:
      dest = os.getcwd() + "/"
    with open(f"{dest}{self.project_name}_Injection_Sequence{tail}.csv", "w") as file:
      file.writelines(line1)
      file.writelines(line2)
      file.writelines(data)
    
    print(f"Wrote {self.project_name}_{tail}.csv using -{self.tray}{self.start_index} -{self.tray}{self.pool_well} ")
    #print(f"Wrote {self.project_name}_Injection_Sequence{tail}.csv to {dest} using {self.tray}{self.positions[0]} and {self.pool_well} as Pool")
    del(data)






# 
# file = "Z:\Aaron\IDeA_automate\ChoiE_033123_SampleList.xlsx"
# a = SampleList(file)
# 
# (a.dirname, a.sample_list_name, tray, startindex, a.get_sample_name_list())
# InjectionSequence("Z:\Aaron\IDeA_automate\ChoiE_033123_SampleList.xlsx", "ASDF", "SDF").write
