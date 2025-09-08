import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
from datamodel import DataModel

from SequenceWidgets import TrayPicker

#################################################################
#############################################################
#####    Helper to keep popups centered and looking nice
#############################################################
#################################################################


def show_modal_centered(win: tk.Toplevel, master: tk.Misc, *, topmost_pulse: bool = True) -> None:
    """
    Make a Toplevel modal, raise it above the parent, and center it over the parent.
    Call this after you've created and laid out the dialog's widgets.
    """
    # Tie to parent and make modal
    win.withdraw()
    win.transient(master)
    try:
        win.grab_set()
    except tk.TclError:
        pass  # safe fallback if grabs aren't allowed in some contexts

    # Ensure sizes are computed

    master.update_idletasks()

    # Compute target rect (parent if mapped, else screen)
    try:
        parent_is_visible = bool(master.winfo_viewable())
    except tk.TclError:
        parent_is_visible = False

    if parent_is_visible:
        px, py = master.winfo_rootx(), master.winfo_rooty()
        pw, ph = master.winfo_width(), master.winfo_height()
    else:
        px, py = 0, 0
        pw, ph = win.winfo_screenwidth(), win.winfo_screenheight()

    win.update_idletasks()
    ww, wh = win.winfo_reqwidth(), win.winfo_reqheight()
    x = px + max(0, (pw - ww) // 2)
    y = py + max(0, (ph - wh) // 2)

    # Clamp to screen so it never spawns off-screen
    sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()
    x = max(0, min(x, sw - ww))
    y = max(0, min(y, sh - wh))

    win.geometry(f"+{x}+{y}")
    win.deiconify()
    win.lift()
    if topmost_pulse:
        # Briefly set topmost to ensure it’s in front, then release (avoids sticking above everything forever)
        win.attributes("-topmost", True)
        win.after(50, lambda: win.attributes("-topmost", False))

    # Focus it
    try:
        win.focus_force()
    except tk.TclError:
        pass


class EditMethodsFrame(tk.Frame):
    def __init__(self, parent, key, methods):
        super().__init__(parent)
        self.parent = parent
        self.key = key
        self.methods = methods
        self.var = tk.StringVar(value=self.methods)

        self.label_text = "DDA" if key == 'DDA' else "DIA" if key == 'DIA' else "ERROR"

        self.listFrame = tk.Frame(self)
        self.listFrame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.buttonFrame = tk.Frame(self)
        self.buttonFrame.pack(side=tk.LEFT, fill=tk.Y)

        self.listbox = tk.Listbox(self.listFrame, listvariable=self.var, selectmode=tk.SINGLE, height=10, width=130)
        self.listbox.pack()

        self.addButton = tk.Button(self.buttonFrame, text="Add", command=self.addMethod)
        self.addButton.pack(fill=tk.X)

        self.removeButton = tk.Button(self.buttonFrame, text="Remove", command=self.removeMethod)
        self.removeButton.pack(fill=tk.X)  

        self.upButton = tk.Button(self.buttonFrame, text="Up", command=self.moveUp)
        self.upButton.pack(fill=tk.X)

        self.downButton = tk.Button(self.buttonFrame, text="Down", command=self.moveDown)
        self.downButton.pack(fill=tk.X)    

        return

    def addMethod(self):
        method_path = filedialog.askopenfilename(parent=self, title="Select Method File", filetypes=[("Method Files", "*.meth"), ("All Files", "*.*")])
        if method_path:
            self.methods.append(method_path)
            self.var.set(self.methods)
        return
    
    def removeMethod(self):
        sel = self.listbox.curselection()
        if sel:
            index = sel[0]
            del self.methods[index]
            self.var.set(self.methods)
        return
    
    def moveUp(self):
        sel = self.listbox.curselection()
        if sel:
            index = sel[0]
            if index > 0:
                self.methods[index], self.methods[index - 1] = self.methods[index - 1], self.methods[index]
                self.var.set(self.methods)
                self.listbox.selection_set(index - 1)
        return  
    
    def moveDown(self):
        sel = self.listbox.curselection()
        if sel:
            index = sel[0]
            if index < len(self.methods) - 1:
                self.methods[index], self.methods[index + 1] = self.methods[index + 1], self.methods[index]
                self.var.set(self.methods)
                self.listbox.selection_set(index + 1)
        return  



class SettingsDialog(tk.Toplevel):
    def __init__(self, parent: tk.Misc, datamodel: DataModel):
        super().__init__(parent)
        self.parent = parent
        self.datamodel = datamodel
        self.title("Settings")

        self.settings = datamodel.settings

        self.container = tk.Frame(self)
        self.container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        ####   Default Tray

        self.default_tray_frame = tk.Frame(self.container)   
        self.default_tray_frame.pack(side=tk.TOP, fill=tk.X, pady=5)

        self.default_tray_label = tk.Label(self.default_tray_frame, text="Default Tray:")
        self.default_tray_label.pack(side=tk.TOP)

        self.default_tray_var = tk.StringVar(value=self.settings.default_tray)
        self.default_tray_picker = TrayPicker(self.default_tray_frame, textvariable=self.default_tray_var, onChange=None, initial_val=self.settings.default_tray)     
        self.default_tray_picker.pack(side=tk.TOP)

        ####  Default Instrument

        self.default_instrument_frame = tk.Frame(self.container)   
        self.default_instrument_frame.pack(side=tk.TOP, fill=tk.X, pady=5)

        self.default_instrument_label = tk.Label(self.default_instrument_frame, text="Default Instrument:")
        self.default_instrument_label.pack(side=tk.TOP)

        self.default_instrument_var = tk.StringVar(value=self.settings.default_instrument)
        self.default_instrument_combo = ttk.Combobox(self.default_instrument_frame, textvariable=self.default_instrument_var, values=self.datamodel.instrument_list, state='readonly', width=20)
        self.default_instrument_combo.pack(side=tk.TOP)

        ###  Buttons

        self.button_frame = tk.Frame(self.container)
        self.button_frame.pack(side=tk.TOP, fill=tk.X, pady=10)

        self.applyButton = tk.Button(self.button_frame, text="Apply", command=self.onDone)
        self.applyButton.pack(side=tk.LEFT, padx=5)

        self.cancelButton = tk.Button(self.button_frame, text="Cancel", command=self.onCancel)
        self.cancelButton.pack(side=tk.LEFT, padx=5)

        show_modal_centered(self, parent)
        self.wait_window()

        return
    
    def onDone(self):
        self.settings.default_tray = self.default_tray_var.get()
        self.settings.default_instrument = self.default_instrument_var.get()
        self.datamodel.settings.saveToFile()
        self.destroy()
        return
    
    def onCancel(self):
        self.destroy()
        return



class InstrumentSettingsDialog(tk.Toplevel):
    def __init__(self, parent: tk.Misc, datamodel: DataModel):
        super().__init__(parent)
        self.parent = parent
        self.datamodel = datamodel
        self.title("Instrument Settings")

        self.current_settings = self.datamodel.instrument_data

        self.container = tk.Frame(self)
        self.container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.instrument_frame = tk.Frame(self.container)
        self.instrument_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        ####   Instrument Selection
        instrument_list = list(self.current_settings.keys())
        self.instrument_var = tk.StringVar(value=self.datamodel.getOption('instrument'))
        self.instrument_label = tk.Label(self.instrument_frame, text="Select Instrument:")
        self.instrument_label.grid(row=0, column=0, sticky=tk.W, pady=5)
        self.instrument_combo = ttk.Combobox(self.instrument_frame, textvariable=self.instrument_var, values=instrument_list, state='readonly', width=30)
        self.instrument_combo.grid(row=0, column=1, sticky=tk.W, pady=5)
        self.instrument_combo.bind("<<ComboboxSelected>>", self.onInstrumentChange)

        self.settings_frame = tk.Frame(self.container)
        self.settings_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        self.buildSettingsFrame()

        self.buttons_frame = tk.Frame(self.container)
        self.buttons_frame.pack(side=tk.TOP, fill=tk.X, pady=10)

        self.apply_all_button = tk.Button(self.buttons_frame, text="Apply Forever", command=self.applyAll)
        self.apply_all_button.pack(side=tk.LEFT, padx=5)

        self.apply_button = tk.Button(self.buttons_frame, text="Apply Current", command=self.applyCurrent)
        self.apply_button.pack(side=tk.LEFT, padx=5)

        self.cancel_button = tk.Button(self.buttons_frame, text="Cancel", command=self.onCancel)
        self.cancel_button.pack(side=tk.LEFT, padx=5)

        show_modal_centered(self, parent)
        self.wait_window()

        return
    
    def onInstrumentChange(self, event):
        self.buildSettingsFrame()
        return
    
    def applyCurrent(self):
        instrument = self.instrument_var.get()
        if instrument in self.current_settings:
            settings = self.current_settings[instrument]

            settings['methods']['DDA'] = self.dda_methods
            settings['methods']['DIA'] = self.dia_methods
            settings['methods']['QC'] = self.qc_var.get()
            settings['methods']['blank'] = self.blank_var.get()
            settings['methods']['rinse'] = self.rinse_var.get()
            settings['methods']['end'] = self.end_var.get()

            # Save back to datamodel
            self.datamodel.instrument_data[instrument] = settings
        self.destroy()
        return
    
    def applyAll(self):
        self.applyCurrent()
        self.datamodel.save_instrument_data()
        # self.destroy()
        return
    
    def onCancel(self):
        self.destroy()
        return
    
    def buildSettingsFrame(self):
        for widget in self.settings_frame.winfo_children():
            widget.destroy()

        instrument = self.instrument_var.get()
        if instrument in self.current_settings:
            settings = self.current_settings[instrument]

        self.dda_label = tk.Label(self.settings_frame, text="DDA Methods:")
        self.dda_label.pack(fill=tk.X, pady=5)
        self.dda_methods = settings['methods']['DDA']
        # self.dda_var = tk.StringVar(value=self.dda_methods)
        self.dda_selector = EditMethodsFrame(self.settings_frame, "DDA", self.dda_methods)
        self.dda_selector.pack(fill=tk.X, pady=5)

        
        self.dia_label = tk.Label(self.settings_frame, text="DIA Methods:")
        self.dia_label.pack(fill=tk.X, pady=5)
        self.dia_methods = settings['methods']['DIA']
        # self.dia_var = tk.StringVar(value=self.dia_methods)
        self.dia_selector = EditMethodsFrame(self.settings_frame, "DIA", self.dia_methods)
        self.dia_selector.pack(fill=tk.X, pady=5)

        self.qc_var = tk.StringVar(value=settings['methods']['QC'])
        self.qc_selector = self.methodSelector(self.settings_frame, "QC Method:", 'QC', self.qc_var)
        self.qc_selector.pack(fill=tk.X, pady=5)

        self.blank_var = tk.StringVar(value=settings['methods']['blank'])
        self.blank_selector = self.methodSelector(self.settings_frame, "Blank Method:", 'blank', self.blank_var)
        self.blank_selector.pack(fill=tk.X, pady=5)

        self.rinse_var = tk.StringVar(value=settings['methods']['rinse'])
        self.rinse_selector = self.methodSelector(self.settings_frame, "Rinse Method:", 'rinse', self.rinse_var)
        self.rinse_selector.pack(fill=tk.X, pady=5)  

        self.end_var = tk.StringVar(value=settings['methods']['end'])
        self.end_selector = self.methodSelector(self.settings_frame, "End Method:", 'end', self.end_var)
        self.end_selector.pack(fill=tk.X, pady=5)

        return

    
    def methodSelector(self, parent, label_text, key, var):

        ####  sample methods go by a different widget
        if key == 'DDA' or key == 'DIA':
            return

        frame = tk.Frame(parent)
        frame.pack(fill=tk.X, pady=5)

        label = tk.Label(frame, text=label_text)
        label.pack(side=tk.LEFT)

        entry = tk.Entry(frame, textvariable=var, width=130)
        entry.pack(side=tk.LEFT, padx=5)

        def onBrowse():
            method_path = filedialog.askopenfilename(parent=self, title="Select Method File", filetypes=[("Method Files", "*.meth"), ("All Files", "*.*")])
            if method_path:
                var.set(method_path)
            return

        button = tk.Button(frame, text="Browse", command=onBrowse)
        button.pack(side=tk.LEFT)

        return frame



# class EditMethodsDialog(tk.Toplevel):
#     def __init__(self, parent: tk.Misc, datamodel: DataModel):
#         super().__init__(parent)
#         self.parent = parent
#         self.datamodel = datamodel
#         self.title("Edit Method List")

#         self.result = None  # Will hold the updated list of methods or None if cancelled

#         ###  A list of the methods available for the selected instrument
#         self.methods = self.datamodel.getInstrumentData('methods')[self.datamodel.getOption('diadda')]

#         self.methodsVar = tk.StringVar(value=self.methods)

#         self.container = tk.Frame(self)
#         self.container.pack(fill=tk.BOTH, expand=True)

#         self.listFrame = tk.Frame(self.container)
#         self.listFrame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

#         self.buttonFrame = tk.Frame(self.container)
#         self.buttonFrame.pack(side=tk.LEFT, fill=tk.Y)

#         self.listbox = tk.Listbox(self.listFrame, listvariable=self.methodsVar, selectmode=tk.SINGLE, height=10, width=50)
#         self.listbox.pack()

#         self.addButton = tk.Button(self.buttonFrame, text="Add", command=self.addMethod)
#         self.addButton.pack(fill=tk.X)

#         self.removeButton = tk.Button(self.buttonFrame, text="Remove", command=self.removeMethod)
#         self.removeButton.pack(fill=tk.X)  

#         self.upButton = tk.Button(self.buttonFrame, text="Up", command=self.moveUp)
#         self.upButton.pack(fill=tk.X)

#         self.downButton = tk.Button(self.buttonFrame, text="Down", command=self.moveDown)
#         self.downButton.pack(fill=tk.X)      

#         self.doneButton = tk.Button(self.buttonFrame, text="Done", command=self.onDone)
#         self.doneButton.pack(fill=tk.X) 

#         self.cancelButton = tk.Button(self.buttonFrame, text="Cancel", command=self.onCancel)
#         self.cancelButton.pack(fill=tk.X)   

#         show_modal_centered(self, parent)
#         self.wait_window()

#         return
    
#     def addMethod(self):
#         method_path = filedialog.askopenfilename(parent=self, title="Select Method File", filetypes=[("Method Files", "*.meth"), ("All Files", "*.*")])
#         if method_path:
#             self.methods.append(method_path)
#             self.methodsVar.set(self.methods)
#         return
    
#     def removeMethod(self):
#         sel = self.listbox.curselection()
#         if sel:
#             index = sel[0]
#             del self.methods[index]
#             self.methodsVar.set(self.methods)
#         return
    
#     def moveUp(self):
#         sel = self.listbox.curselection()
#         if sel:
#             index = sel[0]
#             if index > 0:
#                 self.methods[index], self.methods[index - 1] = self.methods[index - 1], self.methods[index]
#                 self.methodsVar.set(self.methods)
#                 self.listbox.selection_set(index - 1)
#         return  
    
#     def moveDown(self):
#         sel = self.listbox.curselection()
#         if sel:
#             index = sel[0]
#             if index < len(self.methods) - 1:
#                 self.methods[index], self.methods[index + 1] = self.methods[index + 1], self.methods[index]
#                 self.methodsVar.set(self.methods)
#                 self.listbox.selection_set(index + 1)
#         return  
    
#     def onDone(self):
#         self.result = self.methods
#         self.destroy()
#         return
    
#     def onCancel(self):
#         self.result = None
#         self.destroy()
#         return

#########
        ##   Old Exception Handlers from the old version for dealing with Sample List issues


# class AskHeaderRow(tk.Toplevel):

#     def __init__(self, parent, sh):

#         self.parent = parent
#         tk.Toplevel.__init__(self, self.parent)

#         self.sh = sh

#         self.choice = tk.StringVar(value="None")

#         self.top_frame = tk.Frame(self)
#         self.label = tk.Label(self.top_frame, text="Could not find header row.\n  Please select the header for sample number.")
#         self.label.pack()
#         self.top_frame.pack()

#         self.row_frame = tk.Frame(self)
#         self.row_frame.pack()

#         self.button_frame = tk.Frame(self)
#         self.button_frame.pack()

#         self.row = 0

#         self.displayRow()

#         self.done_button = tk.Button(self.button_frame, text="Done", command=self.onDone)
#         self.done_button.pack(side=tk.LEFT)
#         self.next_button = tk.Button(self.button_frame, text="Next Row", command=self.displayRow)
#         self.next_button.pack(side=tk.LEFT)

#         self.parent.eval(f'tk::PlaceWindow {str(self)} center')
#         self.transient(self.parent)
#         self.grab_set()
#         self.parent.wait_window(self)

#         return

#     def onDone(self):
#         self.destroy()

#     def displayRow(self):
#         for w in self.row_frame.winfo_children():
#             w.destroy()
#         self.row += 1

#         for i in range(1, self.sh.max_column + 1):
#             name = self.sh.cell(row = self.row, column = i).value
#             rad = tk.Radiobutton(self.row_frame, text=name, var=self.choice, value=name)
#             rad.pack()


# class askNameColumn(tk.Toplevel):

#     def __init__(self, parent, sh, row):

#         self.parent = parent
#         tk.Toplevel.__init__(self, self.parent)

#         self.sh = sh

#         self.choice = tk.StringVar(value="None")

#         self.top_frame = tk.Frame(self)
#         self.label = tk.Label(self.top_frame, text="Please select header for Sample Name")
#         self.label.pack()
#         self.top_frame.pack()

#         self.row_frame = tk.Frame(self)
#         self.row_frame.pack()

#         self.button_frame = tk.Frame(self)
#         self.button_frame.pack()

#         self.row = row

#         self.displayRow()

#         self.done_button = tk.Button(self.button_frame, text="Done", command=self.onDone)
#         self.done_button.pack(side=tk.LEFT)

#         self.parent.eval(f'tk::PlaceWindow {str(self)} center')
#         self.transient(self.parent)
#         self.grab_set()
#         self.parent.wait_window(self)

#         return

#     def onDone(self):
#         self.destroy()

#     def displayRow(self):

#         for i in range(1, self.sh.max_column + 1):
#             name = self.sh.cell(row = self.row, column = i).value
#             rad = tk.Radiobutton(self.row_frame, text=name, var=self.choice, value=name)
#             rad.pack()




# class AskProjName(tk.Toplevel):

#     def __init__(self, parent):
#         self.parent = parent
#         tk.Toplevel.__init__(self, self.parent)

#         self.frame = tk.Frame(self)

#         self.label = tk.Label(self.frame, text="Please Enter a Project Name")
#         self.label.pack()

#         self.entry = tk.Entry(self.frame, width=50)
#         self.entry.pack()

#         self.button = tk.Button(self.frame, text="OK", command=self.onOk)
#         self.button.pack()


#         self.frame.pack()

#         self.parent.eval(f'tk::PlaceWindow {str(self)} center')
#         self.transient(self.parent)
#         self.grab_set()
#         self.parent.wait_window(self)

#         return

#     def onOk(self):
#         self.result = self.entry.get()
#         self.destroy()