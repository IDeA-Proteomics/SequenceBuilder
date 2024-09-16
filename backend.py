import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
import re
import json

import frontend
import datamodel

class SampleList():

    def sanitizeName(self, name):
        newName = re.sub(r"[ +\[\]\.\+!@#\$%\^&\*\(\)\?\|\\]+", '_', str(name))
        if newName[0].isdigit():
            newName = 'S' + newName
        return newName


    def __init__(self, front, filename):
        with open("instrument_data.json") as jf:
            datamodel.instrument_data = json.load(jf)
        self.front = front
        self.path = filename

        self.abs_path = os.path.abspath(self.path)

        split_name = self.abs_path.split('\\')[-1]
        proj_pattern = r'([^/]+_\d{6}.*)(?=_SampleList.xlsx)'
        match = re.match(proj_pattern, split_name)

        self.project_name = match.group(1) if match else self.front.askProjectName()

        ####  Examine XLS file.  Find row with headers and last sample (first empty after samples)

        wb = load_workbook(self.abs_path, data_only=True)
        sh = wb.worksheets[0]

        namestr = 'sample name'
        numstr = 'sample number'

        self.head_row = None
        self.last_row = None
        for i in range(1, sh.max_row + 1):
            if self.head_row == None:
                for j in range (1, sh.max_column + 1):
                    if sh.cell(row = i, column=j).value == "sample number":
                        self.head_row = i
                        break
        

        if self.head_row == None:
            headrv = self.front.askHeaderRow(sh)
            self.head_row = headrv[0]
            numstr = headrv[1]

        if namestr not in [cell.value for cell in sh[self.head_row]]:
            namestr = self.front.askNameHeader(sh, self.head_row)

        for i in range(self.head_row, sh.max_row + 1):
            if all([cell.value is None for cell in sh[i]]):
                self.last_row = i
                break
        if self.last_row == None:
            self.last_row = sh.max_row + 1
                



        self.data = pd.read_excel(self.abs_path, engine='openpyxl', skiprows=self.head_row - 1, nrows=(self.last_row - self.head_row))
        self.list = pd.DataFrame(data=self.data[[numstr, namestr]].values, columns=['number', 'name'])
        self.list['method'] = "None"
        self.list['position'] = "NA"
        self.list['name'] = self.list['name'].apply(self.sanitizeName)

        for i in self.list.index:
            name = self.list.loc[i, 'name']
            if type(self.list.loc[i, 'number']) == int:
                name = name + "_sample_{:02d}".format(self.list.loc[i, 'number'])
            else:
                name = name + "_sample_{}".format(self.list.loc[i, 'number'])
            self.list.at[i, 'name'] = name



        self.sequence = pd.DataFrame(
            columns=[
                'Sample Type' , 
                'File Name' , 
                'Sample ID' , 
                'Path' , 
                'Instrument Method' , 
                'Process Method' , 
                'Calibration File' , 
                'Position' , 
                'Inj Vol' , 
                'Level' , 
                'Sample Wt' , 
                'Sample Vol' , 
                'ISTD Amt' , 
                'Dil Factor' , 
                'L1 Study' , 
                'L2 Client' , 
                'L3 Laboratory' , 
                'L4 Company' , 
                'L5 Phone' , 
                'Comment' , 
                'Sample Name'
                ],
                index=(range(self.getSampleCount()))
            )

        return
    
    def buildGPF(self):        
        w = []
        for i in range(1,7):
            mz_start = 300 + (100*i)
            mz_end = 300 + (100*i)
            filename = f'{self.project_name}_GPF6_{i}'
            gpf_method_template = r"C:\\Xcalibur\\methods\\DIA\\60min_5ulLoad_10ulLoop\\DIA_GPF_{mz_start}_{mz_end}_60min_5ulLoad_10ulLoop_031221"
            meth_name = gpf_method_template.format(mz_start=mz_start, mz_end=mz_end)
            w.append({'Sample Type':'Unknown', 
                      'File Name':filename, 
                      'Sample ID':'GPF_' + str(i), 
                      'Path':"D:\\" + self.project_name, 
                      'Instrument Method':meth_name,
                      'Position': self.getPoolWell(),
                      'Inj Vol': str(datamodel.instrument_data[self.getInstrument()]['loop_vol']),
                      'Sample Name':"GPF_" + str(i)
                      })
        df = pd.DataFrame.from_records(w,index=range(6))
        return df
    
    def buildPool(self):
        w = []
        for i in range(1, 4):
            filename = f'{self.project_name}_Pool_{i}'
            w.append({
                'Sample Type':'Unknown',
                'File Name' : filename,
                'Sample ID':'Pool' + str(i),
                'Path':"D:\\" + self.project_name,
                'Instrument Method': self.getMethod(),
                'Position': self.getPoolWell(),
                'Inj Vol': str(datamodel.instrument_data[self.getInstrument()]['loop_vol']),
                'Sample Name':"Pool_" + str(i)
            })
        df = pd.DataFrame.from_records(w, index=range(3))
        return df
    
    def buildBlankQC(self, whichOne):
        blank = {
                'Sample Type':'Unknown',
                'File Name' : "Blank",
                'Sample ID':"Blank",
                'Path':"D:\\" + self.project_name,
                'Instrument Method': datamodel.instrument_data[self.getInstrument()]['methods']['blank'],
                'Position': "G1",
                'Inj Vol': str(datamodel.instrument_data[self.getInstrument()]['loop_vol']),
                'Sample Name':"Blank"
        }
        rinse = {
                'Sample Type':'Unknown',
                'File Name' : "Rinse",
                'Sample ID':"Rinse",
                'Path':"D:\\" + self.project_name,
                'Instrument Method': datamodel.instrument_data[self.getInstrument()]['methods']['rinse'],
                'Position': "G1",
                'Inj Vol': str(datamodel.instrument_data[self.getInstrument()]['loop_vol']),
                'Sample Name':"Rinse"
        }
        qc = {
                'Sample Type':'Unknown',
                'File Name' : "QC_"+ datamodel.instrument_data[self.getInstrument()]['name'] +"_JJN3_iRT_" + ("pre_" if whichOne == "pre" else "post_") + self.project_name + "_DDA",
                'Sample ID':"QC",
                'Path':"D:\\QC",
                'Instrument Method': datamodel.instrument_data[self.getInstrument()]['methods']['QC'],
                'Position': "RA1",
                'Inj Vol': str(datamodel.instrument_data[self.getInstrument()]['loop_vol']),
                'Sample Name':"QC_" + ("PRE" if whichOne == "pre" else "POST")
        }
        end = {
                'Sample Type':'Unknown',
                'File Name' : "End",
                'Sample ID':"End",
                'Path': "D:\\" + self.project_name,
                'Instrument Method': datamodel.instrument_data[self.getInstrument()]['methods']['end'],
                'Position': "G1",
                'Inj Vol': str(datamodel.instrument_data[self.getInstrument()]['loop_vol']),
                'Sample Name':"End"
        }

        count = 5
        qc_list = [rinse, blank, qc, rinse, blank]
        if whichOne != "pre":
            qc_list.append(end)
            count = 6

        return pd.DataFrame.from_records(qc_list, index=range(count))

    def reBuildSequence(self):
        self.sequence = pd.DataFrame(
            columns=[
                'Sample Type' , 
                'File Name' , 
                'Sample ID' , 
                'Path' , 
                'Instrument Method' , 
                'Process Method' , 
                'Calibration File' , 
                'Position' , 
                'Inj Vol' , 
                'Level' , 
                'Sample Wt' , 
                'Sample Vol' , 
                'ISTD Amt' , 
                'Dil Factor' , 
                'L1 Study' , 
                'L2 Client' , 
                'L3 Laboratory' , 
                'L4 Company' , 
                'L5 Phone' , 
                'Comment' , 
                'Sample Name'
                ],
                index=(range(self.getSampleCount()))
            )
        if self.front.getRandom() == 1:
            temp_list = self.list.sample(frac=1).reset_index(drop=True)
        else:
            temp_list = self.list
        self.sequence['Sample Type'] = "Unknown"
        self.sequence['File Name'] = [self.project_name + '_' + str(x) for x in temp_list['name']]
        self.sequence['Sample ID'] = temp_list['number']
        self.sequence['Path'] = "D:\\" + self.project_name 
        self.sequence['Instrument Method'] = self.getMethod()
        self.sequence['Position'] = temp_list['position']
        self.sequence['Inj Vol'] = str(datamodel.instrument_data[self.getInstrument()]['loop_vol'])

        self.sequence['Sample Name'] = temp_list['name']   

        if self.front.getPool() == 1:
            oneThird = int(len(self.sequence)/3)
            pools = self.buildPool()
            self.sequence = pd.concat([
                self.sequence.iloc[:oneThird], pools.iloc[:1],
                self.sequence.iloc[oneThird:oneThird*2], pools.iloc[1:2],
                self.sequence.iloc[oneThird*2:], pools.iloc[2:]
            ]).reset_index(drop=True)
                 
        
        if self.front.getGPF() == 1:
            halfway = int(len(self.sequence)/2)
            self.sequence = pd.concat([self.sequence.iloc[:halfway], self.buildGPF(), self.sequence.iloc[halfway:]]).reset_index(drop=True)

        if (self.getAddBlanks() > 0):
            bl_list = []
            last_i = 0
            for i in range(self.getAddBlanks(), len(self.sequence), self.getAddBlanks()):
                bl_list.append(self.sequence.iloc[last_i : i])
                last_i = i
                blank = {
                'Sample Type':'Unknown',
                'File Name' : "Blank",
                'Sample ID':"Blank",
                'Path':"D:\\" + self.project_name,
                'Instrument Method': datamodel.instrument_data[self.getInstrument()]['methods']['blank'],
                'Position': "G1",
                'Inj Vol': str(datamodel.instrument_data[self.getInstrument()]['loop_vol']),
                'Sample Name':"Blank"
                }
                rinse = {
                'Sample Type':'Unknown',
                'File Name' : "Rinse",
                'Sample ID':"Rinse",
                'Path':"D:\\" + self.project_name,
                'Instrument Method': datamodel.instrument_data[self.getInstrument()]['methods']['rinse'],
                'Position': "G1",
                'Inj Vol': str(datamodel.instrument_data[self.getInstrument()]['loop_vol']),
                'Sample Name':"Rinse"
                }

                bl_list.append(pd.DataFrame.from_records([rinse, blank], index=range(2)))
            
            bl_list.append(self.sequence.iloc[last_i:])

            self.sequence = pd.concat(bl_list).reset_index(drop=True)

        if self.front.getAddQC() == 1:
            self.sequence = pd.concat([self.buildBlankQC("pre"), self.sequence, self.buildBlankQC("post")]).reset_index(drop=True)

        

        return
    
    def outputSequence(self):
        self.reBuildSequence()
        template = "Bracket Type=4,\n"
        fname = "{}\\{}_Injection_Sequence.csv".format(os.path.dirname(self.abs_path), self.project_name)
        with open(fname, 'w') as fp:
            fp.write(template)
        self.sequence.to_csv(fname, index=False, mode='a')

        
        # self.buildGPF()
        return
    
    
    def reBuildList(self):

        self.list['position'] = self.front.getSamplePositions()

        return
    
    def getAddBlanks(self):
        return int(self.front.getAddBlanks())

    def getMethod(self):
        return self.front.getMethod()
    
    def getInstrument(self):
        return self.front.getInstrument()
    
    def getPoolWell(self):
        return self.front.getPoolWell()
    
    def getSampleCount(self):
        return len(self.data.index)
    
    def getSequenceCount(self):
        return len(self.sequence)

    def getListData(self, data, index):
        return self.list[data][index]
     
    def getSequenceData(self, data, index):
        return self.sequence[data][index]
    
    def getSampleName(self, index):
        if index < self.getSampleCount():
            return self.list['sample name'][index]
        else:
            return ""
        
    def getSampleNumber(self, index):
        if index < self.getSampleCount():
            return self.list['sample number'][index]
        else:
            return 'X'
        

if (__name__ == "__main__"):
    frontend.showSeqFE()
        
    
        
    
